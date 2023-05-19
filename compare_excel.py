import openpyxl
from datetime import datetime

# 打开Excel文档
workbook = openpyxl.load_workbook('./对比.xlsx')

# 获取第一页、第二页、第三页
sheet1 = workbook.worksheets[0]
sheet2 = workbook.worksheets[1]
sheet3 = workbook.worksheets[2]

start_time = datetime.now()
end_time = datetime.now()

# 遍历第一页中的每一行，
sheet1_Col_A = sheet1["A"]
sheet1_Col_B = sheet1["B"]

sheet2_Col_A = sheet2["A"]
sheet2_Col_B = sheet2["B"]
for i in range(1, sheet1.max_row):
  a_value = sheet1_Col_A[i].value # 获取a表列a的值
  b_value =  sheet1_Col_B[i].value # 获取a表列b的值
  start_time = datetime.now()
  found_match = False
  for j in range(1, sheet2.max_row):
    sheet2_a_value = sheet2_Col_A[j].value # 获取b表列a的值
    sheet2_b_value = sheet2_Col_B[j].value # 获取b表列b的值
    # print("当前b表:" + str(j))
    if sheet2_a_value == a_value:
      found_match = True
      if sheet2_b_value != b_value:
        sheet3.append(["表1", i+1, a_value, f"{b_value}, {sheet2_b_value}"])

  if not found_match:
    sheet3.append(["表1", i+1, a_value, b_value])

  end_time = datetime.now()
  time_diff = end_time - start_time
  milliseconds = int(time_diff.total_seconds() * 1000)
  print("当前a表:" + str(i)+",此操作花费的毫秒数:"+str(milliseconds))

# 循环遍历第二页每一行列a，依次去跟第一页列a比较
for i in range(1, sheet2.max_row):
  a_value = sheet2_Col_A[i].value # 获取列a的值
  b_value = sheet2_Col_B[i].value # 获取列b的值

  start_time = datetime.now()
  found_match = False
  for cell in sheet1_Col_A:
    # print("当前a表:" + str(j))
    if cell.value == a_value:
      found_match = True
      break

  if not found_match:
    sheet3.append(["表2", i+1, a_value, b_value])

  end_time = datetime.now()
  time_diff = end_time - start_time
  milliseconds = int(time_diff.total_seconds() * 1000)
  print("当前b表:" + str(i) + ",此操作花费的毫秒数:" + str(milliseconds))


# 保存Excel文档
workbook.save('./对比.xlsx')
print('对比完成并已保存到对比页！')